import os
import csv
from openpyxl import load_workbook
from datetime import datetime

def get_all_subjects():
    """Возвращает список всех предметов"""
    return [
        "Математика",
        "Русский язык", 
        "Литература",
        "Иностранный язык",
        "Информатика",
        "История",
        "Обществознание",
        "Физика",
        "Химия",
        "Биология",
        "География",
        "Физическая культура",
        "Основы безопасности жизнедеятельности",
        "Основы профессиональной деятельности"
    ]

def read_students_from_group(group_path):
    """Читает список студентов из файла студенты.xlsx"""
    students_file = os.path.join(group_path, "студенты.xlsx")
    students = []
    
    if os.path.exists(students_file):
        try:
            wb = load_workbook(students_file)
            ws = wb.active
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] and row[2] and row[3]:  # Фамилия, Имя, Отчество
                    fio = f"{row[1]} {row[2]} {row[3]}"
                    students.append(fio)
        except Exception as e:
            print(f"Ошибка при чтении файла студентов: {e}")
    
    return students

def get_student_grades(student_fio, group_path, subjects):
    """Получает все оценки студента по всем предметам"""
    student_data = {"ФИО": student_fio}
    
    for subject in subjects:
        subject_file = os.path.join(group_path, f"{subject}.xlsx")
        
        if os.path.exists(subject_file):
            try:
                wb = load_workbook(subject_file)
                ws = wb.active
                
                # Находим строку студента
                student_row = None
                for r in range(2, ws.max_row + 1):
                    if ws.cell(row=r, column=1).value == student_fio:
                        student_row = r
                        break
                
                if student_row:
                    # Собираем все оценки студента
                    grades = []
                    absences = 0
                    
                    for c in range(3, ws.max_column + 1):  # Начинаем с 3-го столбца
                        cell_value = ws.cell(row=student_row, column=c).value
                        
                        if cell_value and isinstance(cell_value, (int, float)) and 2 <= cell_value <= 5:
                            grades.append(int(cell_value))
                        elif str(cell_value).strip() == "Н":
                            absences += 1
                    
                    # Добавляем данные по предмету
                    student_data[f"{subject}_оценки"] = ";".join(map(str, grades)) if grades else ""
                    student_data[f"{subject}_пропуски"] = absences
                    student_data[f"{subject}_средний_балл"] = round(sum(grades) / len(grades), 2) if grades else 0
                else:
                    # Студент не найден в журнале по предмету
                    student_data[f"{subject}_оценки"] = ""
                    student_data[f"{subject}_пропуски"] = 0
                    student_data[f"{subject}_средний_балл"] = 0
                    
            except Exception as e:
                print(f"Ошибка при обработке предмета {subject} для студента {student_fio}: {e}")
                student_data[f"{subject}_оценки"] = ""
                student_data[f"{subject}_пропуски"] = 0
                student_data[f"{subject}_средний_балл"] = 0
        else:
            # Файл предмета не найден
            student_data[f"{subject}_оценки"] = ""
            student_data[f"{subject}_пропуски"] = 0
            student_data[f"{subject}_средний_балл"] = 0
    
    return student_data

def generate_csv_with_grades():
    """Создает CSV файл с ФИО студентов и всеми их оценками"""
    
    # Создаем папку "Итог" если её нет
    result_folder = "Итог"
    os.makedirs(result_folder, exist_ok=True)
    
    # Путь к папке с журналами
    journals_path = "Журналы/1 Курс"
    
    if not os.path.exists(journals_path):
        print(f"Папка {journals_path} не найдена!")
        return
    
    # Получаем список всех групп
    groups = []
    for item in os.listdir(journals_path):
        if os.path.isdir(os.path.join(journals_path, item)):
            groups.append(item)
    
    print(f"Найдено групп: {len(groups)}")
    
    # Получаем список предметов
    subjects = get_all_subjects()
    
    # Подготавливаем заголовки CSV
    headers = ["Группа", "ФИО"]
    for subject in subjects:
        headers.extend([
            f"{subject}_оценки",
            f"{subject}_пропуски", 
            f"{subject}_средний_балл"
        ])
    
    # Создаем CSV файл
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_filename = os.path.join(result_folder, f"Оценки_студентов_{timestamp}.csv")
    
    with open(csv_filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=headers)
        writer.writeheader()
        
        # Обрабатываем каждую группу
        for group_name in groups:
            print(f"Обрабатываем группу: {group_name}")
            group_path = os.path.join(journals_path, group_name)
            
            # Читаем студентов группы
            students = read_students_from_group(group_path)
            
            # Получаем данные по каждому студенту
            for student_fio in students:
                student_data = get_student_grades(student_fio, group_path, subjects)
                student_data["Группа"] = group_name
                
                # Записываем данные студента в CSV
                writer.writerow(student_data)
    
    print(f"\nCSV файл создан: {csv_filename}")
    print("Готово!")
    
    return csv_filename

def generate_simple_csv_with_grades():
    """Создает упрощенный CSV файл с ФИО и средними баллами по предметам"""
    
    # Создаем папку "Итог" если её нет
    result_folder = "Итог"
    os.makedirs(result_folder, exist_ok=True)
    
    # Путь к папке с журналами
    journals_path = "Журналы/1 Курс"
    
    if not os.path.exists(journals_path):
        print(f"Папка {journals_path} не найдена!")
        return
    
    # Получаем список всех групп
    groups = []
    for item in os.listdir(journals_path):
        if os.path.isdir(os.path.join(journals_path, item)):
            groups.append(item)
    
    print(f"Найдено групп: {len(groups)}")
    
    # Получаем список предметов
    subjects = get_all_subjects()
    
    # Подготавливаем заголовки CSV
    headers = ["Группа", "ФИО"] + subjects + ["Общий_средний_балл", "Общее_количество_пропусков"]
    
    # Создаем CSV файл
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_filename = os.path.join(result_folder, f"Средние_баллы_{timestamp}.csv")
    
    with open(csv_filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=headers)
        writer.writeheader()
        
        # Обрабатываем каждую группу
        for group_name in groups:
            print(f"Обрабатываем группу: {group_name}")
            group_path = os.path.join(journals_path, group_name)
            
            # Читаем студентов группы
            students = read_students_from_group(group_path)
            
            # Получаем данные по каждому студенту
            for student_fio in students:
                student_data = {"Группа": group_name, "ФИО": student_fio}
                
                all_grades = []
                total_absences = 0
                
                for subject in subjects:
                    subject_file = os.path.join(group_path, f"{subject}.xlsx")
                    avg_grade = 0
                    absences = 0
                    
                    if os.path.exists(subject_file):
                        try:
                            wb = load_workbook(subject_file)
                            ws = wb.active
                            
                            # Находим строку студента
                            student_row = None
                            for r in range(2, ws.max_row + 1):
                                if ws.cell(row=r, column=1).value == student_fio:
                                    student_row = r
                                    break
                            
                            if student_row:
                                # Собираем оценки и пропуски
                                grades = []
                                for c in range(3, ws.max_column + 1):
                                    cell_value = ws.cell(row=student_row, column=c).value
                                    
                                    if cell_value and isinstance(cell_value, (int, float)) and 2 <= cell_value <= 5:
                                        grades.append(int(cell_value))
                                    elif str(cell_value).strip() == "Н":
                                        absences += 1
                                
                                if grades:
                                    avg_grade = round(sum(grades) / len(grades), 2)
                                    all_grades.extend(grades)
                                
                                total_absences += absences
                                
                        except Exception as e:
                            print(f"Ошибка при обработке предмета {subject} для студента {student_fio}: {e}")
                    
                    student_data[subject] = avg_grade
                
                # Вычисляем общий средний балл
                overall_avg = round(sum(all_grades) / len(all_grades), 2) if all_grades else 0
                student_data["Общий_средний_балл"] = overall_avg
                student_data["Общее_количество_пропусков"] = total_absences
                
                # Записываем данные студента в CSV
                writer.writerow(student_data)
    
    print(f"\nУпрощенный CSV файл создан: {csv_filename}")
    print("Готово!")
    
    return csv_filename

if __name__ == "__main__":
    print("Выберите тип CSV файла:")
    print("1. Полный файл с детальными оценками")
    print("2. Упрощенный файл со средними баллами")
    
    choice = input("Введите номер (1 или 2): ").strip()
    
    if choice == "1":
        generate_csv_with_grades()
    elif choice == "2":
        generate_simple_csv_with_grades()
    else:
        print("Неверный выбор. Создаем упрощенный файл...")
        generate_simple_csv_with_grades()

