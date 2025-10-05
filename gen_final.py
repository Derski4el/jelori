import os
import random
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from typing import List, Dict, Tuple, Optional
import logging
from difflib import SequenceMatcher
import re

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class MonthlyAssessmentGenerator:
    """Класс для генерации месячной аттестации с оптимизациями"""
    
    # Константы
    SUBJECTS = [
        "Математика", "Русский язык", "Литература", "Иностранный язык",
        "Информатика", "История", "Обществознание", "Физика", "Химия",
        "Биология", "География", "Физическая культура",
        "Основы безопасности жизнедеятельности", "Основы профессиональной деятельности"
    ]
    
    # Стили (создаются один раз)
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    
    SUBJECT_HEADER_FONT = Font(bold=True, color="FFFFFF")
    SUBJECT_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    SUBJECT_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    def __init__(self, journals_path: str = "Журналы/1 Курс", result_folder: str = "Итог"):
        self.journals_path = journals_path
        self.result_folder = result_folder
        self.workbook_cache = {}  # Кэш для открытых файлов
        self.student_data_cache = {}  # Кэш данных студентов
        
        # Создаем папку результатов
        os.makedirs(result_folder, exist_ok=True)

    def get_working_days_for_month(self, year: int, month: int) -> List[str]:
        """Универсальная функция для получения рабочих дней любого месяца"""
        from datetime import datetime, timedelta
        working_days = []
        
        if month == 12:
            next_month = datetime(year + 1, 1, 1)
        else:
            next_month = datetime(year, month + 1, 1)
        
        last_day = (next_month - timedelta(days=1)).day
        
        start_date = datetime(year, month, 1)
        end_date = datetime(year, month, last_day)
        
        current_date = start_date
        while current_date <= end_date:
            if current_date.weekday() < 5:  # 0-4 = пн-пт
                working_days.append(current_date.strftime("%d.%m.%Y"))
            current_date += timedelta(days=1)
        
        return working_days

    def calculate_average_grade(self, grades: List[float]) -> float:
        """Вычисляет средний балл из списка оценок"""
        if not grades:
            return 0
        return round(sum(grades) / len(grades), 0)
    
    def get_groups(self) -> List[str]:
        """Получает список всех групп"""
        if not os.path.exists(self.journals_path):
            logger.error(f"Папка {self.journals_path} не найдена!")
            return []
        
        groups = [item for item in os.listdir(self.journals_path) if os.path.isdir(os.path.join(self.journals_path, item))]
        logger.info(f"Найдено групп: {len(groups)}")
        return groups
    
    def load_workbook_cached(self, file_path: str) -> Optional[Workbook]:
        """Загружает рабочую книгу с кэшированием"""
        if file_path in self.workbook_cache:
            return self.workbook_cache[file_path]
        
        try:
            if os.path.exists(file_path):
                wb = load_workbook(file_path, data_only=True)
                self.workbook_cache[file_path] = wb
                return wb
        except Exception as e:
            logger.error(f"Ошибка при загрузке файла {file_path}: {e}")
        
        return None
    
    def get_students_from_group(self, group_name: str) -> List[str]:
        """Получает список студентов группы с кэшированием"""
        cache_key = f"students_{group_name}"
        if cache_key in self.student_data_cache:
            return self.student_data_cache[cache_key]
        
        students_file = os.path.join(self.journals_path, group_name, "студенты.xlsx")
        students = []
        
        wb = self.load_workbook_cached(students_file)
        if wb:
            try:
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > 3 and row[1] and row[2] and row[3]:
                        fio = f"{row[1]} {row[2]} {row[3]}"
                        students.append(fio)
            except Exception as e:
                logger.error(f"Ошибка при чтении студентов группы {group_name}: {e}")
        
        self.student_data_cache[cache_key] = students
        return students

    def get_student_grades_from_subject(self, group_name: str, subject: str, student_fio: str, target_month: int = None, start_date: datetime = None, end_date: datetime = None) -> Tuple[List[float], int, int]:
        """Получает оценки, пропуски и количество занятий студента по предмету"""
        subject_file = os.path.join(self.journals_path, group_name, f"{subject}.xlsx")
        grades = []
        absences = 0
        lessons_count = 0
        
        wb = self.load_workbook_cached(subject_file)
        if not wb:
            return grades, absences, lessons_count
        
        try:
            ws = wb.active
            
            student_row = None
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=1).value == student_fio:
                    student_row = r
                    break
            
            if student_row:
                columns_to_process = []
                if start_date and end_date:
                    columns_to_process = self._get_date_range_columns(ws, start_date, end_date)
                elif target_month:
                    target_dates = self.get_working_days_for_month(2025, target_month)
                    columns_to_process = self._get_month_columns(ws, target_dates)
                else:
                    columns_to_process = list(range(3, ws.max_column + 1))

                if not columns_to_process:
                    return grades, absences, lessons_count
                
                for c in columns_to_process:
                    if c > ws.max_column:
                        continue
                        
                    grade = ws.cell(row=student_row, column=c).value
                    if grade is None or str(grade).strip() == "":
                        continue
                    
                    lessons_count += 1
                    if isinstance(grade, (int, float)) and 2 <= grade <= 5:
                        grades.append(grade)
                    elif str(grade).strip().upper() == "Н":
                        absences += 1
                        
        except Exception as e:
            logger.error(f"Ошибка при обработке предмета {subject} для студента {student_fio}: {e}")
        
        return grades, absences, lessons_count

    def _get_month_columns(self, ws, target_dates: List[str]) -> List[int]:
        """Находит номера столбцов с датами целевого месяца"""
        month_columns = []
        target_dates_set = set(target_dates)
        
        for col in range(3, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value and isinstance(cell_value, datetime):
                date_str = cell_value.strftime("%d.%m.%Y")
                if date_str in target_dates_set:
                    month_columns.append(col)
            elif cell_value and isinstance(cell_value, str):
                if cell_value.strip() in target_dates_set:
                    month_columns.append(col)
        
        return month_columns

    def _get_date_range_columns(self, ws, start_date: datetime, end_date: datetime) -> List[int]:
        """Находит номера столбцов, даты которых попадают в указанный диапазон."""
        date_columns = []
        for col in range(3, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if not cell_value:
                continue
            
            cell_date = None
            if isinstance(cell_value, datetime):
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    date_str = cell_value.split(' ')[0]
                    cell_date = datetime.strptime(date_str, "%d.%m.%Y")
                except ValueError:
                    continue
            
            if cell_date and start_date <= cell_date <= end_date:
                date_columns.append(col)
        return date_columns

    def _get_month_name(self, month: int) -> str:
        """Возвращает название месяца по номеру"""
        month_names = {9: "сентября", 10: "октября", 11: "ноября", 12: "декабря"}
        return month_names.get(month, "неизвестный")
    
    def apply_header_styles(self, ws, headers: List[str]):
        """Применяет стили к заголовкам"""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
    
    def auto_adjust_column_width(self, ws):
        """Автоматически подгоняет ширину столбцов"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def process_group(self, wb: Workbook, group_name: str, target_month: int = None, start_date: datetime = None, end_date: datetime = None) -> int:
        """Обрабатывает одну группу и возвращает количество студентов"""
        info_str = ""
        if target_month:
            info_str = f" за {self._get_month_name(target_month)}"
        elif start_date and end_date:
            info_str = f" за период с {start_date.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')}"

        logger.info(f"Обрабатываем группу: {group_name}{info_str}")
        
        ws = wb.create_sheet(title=group_name)
        
        headers = ["ФИО"] + self.SUBJECTS + ["Пропуски (часы)"]
        self.apply_header_styles(ws, headers)
        
        students = self.get_students_from_group(group_name)
        if not students:
            logger.warning(f"В группе {group_name} не найдено студентов. Пропускаем.")
            return 0

        failing_students_count = 0
        students_with_one_2 = 0
        students_with_one_3 = 0
        students_with_one_4 = 0
        students_with_one_5 = 0
        students_4_and_5_only = 0
        total_absences_lessons = 0
        total_lessons = 0

        for row_idx, student_fio in enumerate(students, 2):
            ws.cell(row=row_idx, column=1, value=student_fio)
            
            total_absences = 0
            lessons_for_student = 0
            subject_avgs: List[int] = []
            
            for col_idx, subject in enumerate(self.SUBJECTS, 2):
                grades, subject_absences, subject_lessons = self.get_student_grades_from_subject(
                    group_name, subject, student_fio, target_month, start_date, end_date
                )
                
                total_absences += subject_absences
                lessons_for_student += subject_lessons
                avg_grade = self.calculate_average_grade(grades)
                ws.cell(row=row_idx, column=col_idx, value=int(avg_grade) if avg_grade else 0)
                if avg_grade:
                    subject_avgs.append(int(avg_grade))
            
            absences_hours_col = 2 + len(self.SUBJECTS)
            ws.cell(row=row_idx, column=absences_hours_col, value=total_absences * 2)

            total_absences_lessons += total_absences
            total_lessons += lessons_for_student

            if subject_avgs:
                count_2 = sum(1 for g in subject_avgs if g == 2)
                count_3 = sum(1 for g in subject_avgs if g == 3)
                count_4 = sum(1 for g in subject_avgs if g == 4)
                count_5 = sum(1 for g in subject_avgs if g == 5)

                if count_2 > 0:
                    failing_students_count += 1
                if count_2 == 1:
                    students_with_one_2 += 1
                if count_3 == 1:
                    students_with_one_3 += 1
                if count_4 == 1:
                    students_with_one_4 += 1
                if count_5 == 1:
                    students_with_one_5 += 1
                if count_2 == 0 and count_3 == 0 and (count_4 > 0 or count_5 > 0):
                    students_4_and_5_only += 1
        
        current_last_row = 1 + len(students)
        metrics_start_row = current_last_row + 2

        students_count = len(students)
        avg_absences_per_student_hours = (total_absences_lessons * 2) / students_count if students_count > 0 else 0.0
        attendance_percent = 100.0 * (1 - (total_absences_lessons / total_lessons)) if total_lessons > 0 else 0.0
        success_percent = 100.0 * ((students_count - failing_students_count) / students_count) if students_count > 0 else 0.0

        metrics = [
            ("Неуспевающих, чел.", failing_students_count),
            ("Студентов с одной '2', чел.", students_with_one_2),
            ("Студентов с одной '3', чел.", students_with_one_3),
            ("Студентов с одной '4', чел.", students_with_one_4),
            ("Студентов с одной '5', чел.", students_with_one_5),
            ("Кол-во пропусков на 1 студента, часов", round(avg_absences_per_student_hours, 1)),
            ("Посещаемость, %", round(attendance_percent, 1)),
            ("Учатся на 4 и 5, чел.", students_4_and_5_only),
            ("Число студентов, чел.", students_count),
            ("Успеваемость, %", round(success_percent, 1)),
        ]

        for idx, (label, value) in enumerate(metrics, start=0):
            ws.cell(row=metrics_start_row + idx, column=1, value=label)
            ws.cell(row=metrics_start_row + idx, column=2, value=value)

        self.auto_adjust_column_width(ws)
        
        logger.info(f"Лист для группы {group_name} создан ({len(students)} студентов)")
        return len(students)
    
    def cleanup_cache(self):
        """Очищает кэш открытых файлов"""
        for wb in self.workbook_cache.values():
            try:
                wb.close()
            except:
                pass
        self.workbook_cache.clear()
        self.student_data_cache.clear()

    def find_students_by_name(self, search_name: str) -> List[Tuple[str, str, float]]:
        """Ищет студентов по ФИО с учетом неточности ввода"""
        matching_students = []
        groups = self.get_groups()
        for group_name in groups:
            students = self.get_students_from_group(group_name)
            for student_fio in students:
                if search_name.lower() in student_fio.lower():
                    matching_students.append((group_name, student_fio, 1.0))
        return matching_students

    def get_all_student_grades(self, group_name: str, student_fio: str, target_month: int = None, start_date: datetime = None, end_date: datetime = None) -> Dict[str, Dict]:
        """Получает все оценки студента по всем предметам"""
        student_grades = {}
        
        for subject in self.SUBJECTS:
            grades, absences, lessons = self.get_student_grades_from_subject(group_name, subject, student_fio, target_month, start_date, end_date)
            average = self.calculate_average_grade(grades)
            
            student_grades[subject] = {
                'grades': grades,
                'absences': absences,
                'average': average
            }
        
        return student_grades
    
    def display_student_grades(self, group_name: str, student_fio: str, student_grades: Dict[str, Dict]):
        """Выводит на экран все оценки студента"""
        print(f"\nОценки для {student_fio} ({group_name}):")
        for subject, data in student_grades.items():
            print(f"  {subject}: {data['average']}")

    def search_and_display_student(self, search_name: str):
        """Основная функция для поиска студента и вывода его оценок"""
        print(f"\n[ПОИСК] Студента: '{search_name}'")
        
        # Ищем студентов
        matching_students = self.find_students_by_name(search_name)
        
        if not matching_students:
            print("[ОШИБКА] Студенты не найдены. Попробуйте изменить поисковый запрос.")
            return
        
        print(f"\n[НАЙДЕНО] Студентов: {len(matching_students)}")
        
        # Показываем найденных студентов
        for i, (group_name, student_fio, similarity) in enumerate(matching_students, 1):
            print(f"{i}. {student_fio} (группа: {group_name}, схожесть: {similarity:.2f})")
        
        # Если найден только один студент, сразу показываем его оценки
        if len(matching_students) == 1:
            group_name, student_fio, _ = matching_students[0]
            student_grades = self.get_all_student_grades(group_name, student_fio)
            self.display_student_grades(group_name, student_fio, student_grades)
        else:
            # Если несколько студентов, просим выбрать
            try:
                choice = input(f"\nВыберите номер студента (1-{len(matching_students)}) или 0 для выхода: ")
                choice_num = int(choice)
                
                if choice_num == 0:
                    print("Поиск отменен.")
                    return
                elif 1 <= choice_num <= len(matching_students):
                    group_name, student_fio, _ = matching_students[choice_num - 1]
                    student_grades = self.get_all_student_grades(group_name, student_fio)
                    self.display_student_grades(group_name, student_fio, student_grades)
                else:
                    print("[ОШИБКА] Неверный номер.")
            except ValueError:
                print("[ОШИБКА] Введите корректный номер.", exc_info=True)

    def create_monthly_assessment(self, month: int = None) -> str:
        """Создает итоговую таблицу 'Месячная аттестация'"""
        try:
            groups = self.get_groups()
            if not groups:
                return ""
            
            wb = Workbook()
            wb.remove(wb.active)
            
            month_name = f"_{self._get_month_name(month)}_2025" if month else ""
            logger.info(f"Создаем аттестацию{month_name.replace('_', ' ')}...")
            
            total_students = 0
            for group_name in groups:
                total_students += self.process_group(wb, group_name, target_month=month)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = os.path.join(self.result_folder, f"Месячная аттестация{month_name}_{timestamp}.xlsx")
            wb.save(filename)
            
            logger.info(f"Файл сохранен: {filename}")
            return filename
            
        except Exception as e:
            logger.error(f"Критическая ошибка при создании аттестации: {e}")
            return ""
        finally:
            self.cleanup_cache()

    def create_assessment_for_date_range(self, start_date: datetime, end_date: datetime) -> str:
        """Создает аттестацию за указанный диапазон дат."""
        try:
            groups = self.get_groups()
            if not groups:
                return ""

            wb = Workbook()
            wb.remove(wb.active)

            logger.info(f"Создаем аттестацию за период с {start_date.strftime('%d.%m.%Y')} по {end_date.strftime('%d.%m.%Y')}...")

            total_students = 0
            for group_name in groups:
                total_students += self.process_group(wb, group_name, start_date=start_date, end_date=end_date)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            start_str = start_date.strftime("%Y%m%d")
            end_str = end_date.strftime("%Y%m%d")
            filename = os.path.join(self.result_folder, f"Аттестация_с_{start_str}_по_{end_str}_{timestamp}.xlsx")
            
            wb.save(filename)
            
            logger.info(f"Файл сохранен: {filename}")
            return filename

        except Exception as e:
            logger.error(f"Критическая ошибка при создании аттестации по датам: {e}")
            return ""
        finally:
            self.cleanup_cache()

def main():
    """Основная функция для запуска CLI"""
    generator = MonthlyAssessmentGenerator()
    
    try:
        while True:
            print("\n" + "="*50)
            print("СИСТЕМА ГЕНЕРАЦИИ АТТЕСТАЦИЙ")
            print("="*50)
            print("1. Найти студента и показать оценки")
            print("2. Создать общую аттестацию (по всем данным)")
            print("3. Создать аттестацию за конкретный месяц")
            print("4. Создать аттестации за все месяцы (сентябрь-декабрь)")
            print("5. Аттестация за выбранный период")
            print("0. Выход")
            
            choice = input("\nВведите номер действия: ").strip()
            
            if choice == "1":
                search_name = input("\nВведите ФИО студента (можно не точно): ").strip()
                if search_name:
                    generator.search_and_display_student(search_name)
                else:
                    print("[ОШИБКА] Введите ФИО студента.")
            
            elif choice == "2":
                print("\n[СОЗДАНИЕ] Общей месячной аттестации...")
                result = generator.create_monthly_assessment()
                if result:
                    print(f"[УСПЕХ] Аттестация создана: {result}")
            
            elif choice == "3":
                month_choice = input("Введите номер месяца (9-12): ").strip()
                try:
                    month = int(month_choice)
                    if 9 <= month <= 12:
                        result = generator.create_monthly_assessment(month)
                        if result:
                            print(f"[УСПЕХ] Аттестация создана: {result}")
                    else:
                        print("[ОШИБКА] Неверный номер месяца.")
                except ValueError:
                    print("[ОШИБКА] Введите число.")
            
            elif choice == "4":
                print("\n[СОЗДАНИЕ] Аттестаций за все месяцы...")
                for month in [9, 10, 11, 12]:
                    generator.create_monthly_assessment(month)
                print("[УСПЕХ] Все месячные аттестации созданы.")

            elif choice == "5":
                print("\n[СОЗДАНИЕ] Аттестации за выбранный период...")
                try:
                    start_date_str = input("Введите начальную дату (ДД.ММ.ГГГГ): ").strip()
                    end_date_str = input("Введите конечную дату (ДД.ММ.ГГГГ): ").strip()
                    
                    start_date = datetime.strptime(start_date_str, "%d.%m.%Y")
                    end_date = datetime.strptime(end_date_str, "%d.%m.%Y")
                    
                    if start_date > end_date:
                        print("[ОШИБКА] Начальная дата не может быть позже конечной.")
                    else:
                        result = generator.create_assessment_for_date_range(start_date, end_date)
                        if result:
                            print(f"[УСПЕХ] Аттестация создана: {result}")
                except ValueError:
                    print("[ОШИБКА] Неверный формат даты. Используйте ДД.ММ.ГГГГ.")

            elif choice == "0":
                print("\nДо свидания!")
                break
            
            else:
                print("[ОШИБКА] Неверный выбор. Попробуйте снова.")
    
    except KeyboardInterrupt:
        print("\n\nПрограмма завершена пользователем.")
    except Exception as e:
        logger.error(f"Произошла критическая ошибка: {e}", exc_info=True)
    finally:
        generator.cleanup_cache()

if __name__ == "__main__":
    main()
