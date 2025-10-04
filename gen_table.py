predmety_1_kurs = [
    "Математика",
    "Русский язык",
    "Литература",
    "Иностранный язык",
    "Информатика",
    "История",
    "Обществознание",
    "Физика",
    "Химия",
    "Биология",
    "География",
    "Физическая культура",
    "Основы безопасности жизнедеятельности",
    "Основы профессиональной деятельности",
]


import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def read_students_from_group_file(filename):
    """Читает студентов из файла список_групп.xlsx и возвращает словарь: {группа: [список студентов]}"""
    wb = load_workbook(filename)
    group_students = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        students = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            # row: (№, Фамилия, Имя, Отчество)
            if row[1] and row[2] and row[3]:
                students.append({
                    'фамилия': row[1],
                    'имя': row[2],
                    'отчество': row[3]
                })
        group_students[sheet] = students
    return group_students

def save_students_list(students, folder_path):
    """Сохраняет список студентов в Excel файл"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Список студентов"
    headers = ['№', 'Фамилия', 'Имя', 'Отчество']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    for idx, student in enumerate(students, 1):
        ws.cell(row=idx+1, column=1, value=idx)
        ws.cell(row=idx+1, column=2, value=student['фамилия'])
        ws.cell(row=idx+1, column=3, value=student['имя'])
        ws.cell(row=idx+1, column=4, value=student['отчество'])
    # Автоматическая ширина
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    filename = os.path.join(folder_path, "студенты.xlsx")
    wb.save(filename)

def save_predmet_files(students, folder_path, predmety):
    """Создаёт по одному файлу на каждый предмет с ФИО студентов"""
    for predmet in predmety:
        wb = Workbook()
        ws = wb.active
        ws.title = predmet
        ws.cell(row=1, column=1, value="ФИО")
        ws.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=1, column=1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
        for idx, student in enumerate(students, 1):
            fio = f"{student['фамилия']} {student['имя']} {student['отчество']}"
            ws.cell(row=idx+1, column=1, value=fio)
        ws.column_dimensions['A'].width = 40
        filename = os.path.join(folder_path, f"{predmet}.xlsx")
        wb.save(filename)

def generate_group_folders_with_files_from_group_file(group_file, predmety):
    """Генерирует папки для каждой группы с нужными файлами, используя список студентов из файлa список_групп.xlsx"""
    group_students = read_students_from_group_file(group_file)
    for group_name, students in group_students.items():
        folder_path = os.path.join(os.getcwd(), group_name)
        os.makedirs(folder_path, exist_ok=True)
        save_students_list(students, folder_path)
        save_predmet_files(students, folder_path, predmety)
        print(f"Папка '{group_name}' создана. Список студентов и файлы по предметам сохранены.")

# Пример использования:
# Укажите путь к вашему файлу, например: "список_групп_20240610_153000.xlsx"
group_file = "список_групп.xlsx"

if __name__ == "__main__":
    # Создаём папки "Журналы/1 Курс", если их нет
    base_folder = os.path.join("Журналы", "1 Курс")
    os.makedirs(base_folder, exist_ok=True)
    # Меняем рабочую директорию на "Журналы/1 Курс", чтобы группы создавались там
    os.chdir(base_folder)
    generate_group_folders_with_files_from_group_file(group_file, predmety_1_kurs)
    print("Генерация завершена!")

