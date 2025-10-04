import os
import random
from openpyxl import load_workbook
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment
import re

def get_working_days_september_2025():
    """Возвращает список рабочих дней (пн-пт) сентября 2025 года"""
    working_days = []
    start_date = datetime(2025, 9, 1)  # 1 сентября 2025
    end_date = datetime(2025, 9, 30)   # 30 сентября 2025
    
    current_date = start_date
    while current_date <= end_date:
        # Проверяем, что это рабочий день (понедельник = 0, воскресенье = 6)
        if current_date.weekday() < 5:  # 0-4 = пн-пт
            working_days.append(current_date.strftime("%d.%m.%Y"))
        current_date += timedelta(days=1)
    
    return working_days

def get_working_days_october_2025():
    """Возвращает список рабочих дней (пн-пт) октября 2025 года"""
    working_days = []
    start_date = datetime(2025, 10, 1)  # 1 октября 2025
    end_date = datetime(2025, 10, 31)   # 31 октября 2025
    
    current_date = start_date
    while current_date <= end_date:
        # Проверяем, что это рабочий день (понедельник = 0, воскресенье = 6)
        if current_date.weekday() < 5:  # 0-4 = пн-пт
            working_days.append(current_date.strftime("%d.%m.%Y"))
        current_date += timedelta(days=1)
    
    return working_days

def get_working_days_november_2025():
    """Возвращает список рабочих дней (пн-пт) ноября 2025 года"""
    working_days = []
    start_date = datetime(2025, 11, 1)  # 1 ноября 2025
    end_date = datetime(2025, 11, 30)   # 30 ноября 2025
    
    current_date = start_date
    while current_date <= end_date:
        # Проверяем, что это рабочий день (понедельник = 0, воскресенье = 6)
        if current_date.weekday() < 5:  # 0-4 = пн-пт
            working_days.append(current_date.strftime("%d.%m.%Y"))
        current_date += timedelta(days=1)
    
    return working_days

def get_working_days_december_2025():
    """Возвращает список рабочих дней (пн-пт) декабря 2025 года"""
    working_days = []
    start_date = datetime(2025, 12, 1)  # 1 декабря 2025
    end_date = datetime(2025, 12, 31)   # 31 декабря 2025
    
    current_date = start_date
    while current_date <= end_date:
        # Проверяем, что это рабочий день (понедельник = 0, воскресенье = 6)
        if current_date.weekday() < 5:  # 0-4 = пн-пт
            working_days.append(current_date.strftime("%d.%m.%Y"))
        current_date += timedelta(days=1)
    
    return working_days

def get_working_days_for_month(year, month):
    """Универсальная функция для получения рабочих дней любого месяца"""
    working_days = []
    
    # Определяем последний день месяца
    if month == 12:
        next_month = datetime(year + 1, 1, 1)
    else:
        next_month = datetime(year, month + 1, 1)
    
    last_day = (next_month - timedelta(days=1)).day
    
    start_date = datetime(year, month, 1)
    end_date = datetime(year, month, last_day)
    
    current_date = start_date
    while current_date <= end_date:
        # Проверяем, что это рабочий день (понедельник = 0, воскресенье = 6)
        if current_date.weekday() < 5:  # 0-4 = пн-пт
            working_days.append(current_date.strftime("%d.%m.%Y"))
        current_date += timedelta(days=1)
    
    return working_days

def is_date_string(value):
    """Проверяет, является ли значение датой в формате DD.MM.YYYY"""
    if not value or not isinstance(value, (str, datetime)):
        return False
    
    if isinstance(value, datetime):
        return True
    
    # Проверяем формат DD.MM.YYYY
    date_pattern = r'^\d{2}\.\d{2}\.\d{4}$'
    if re.match(date_pattern, str(value).strip()):
        return True
    
    return False

def get_last_date_column(ws):
    """Находит последний столбец с датой в файле Excel
    
    Returns:
        int: номер последнего столбца с датой, или 2 если даты не найдены
    """
    # Проверяем столбец B (column=2) и далее
    for col in range(2, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if not is_date_string(cell_value):
            return col - 1  # Возвращаем предыдущий столбец
    
    return ws.max_column  # Если все столбцы содержат даты

def has_existing_dates(ws):
    """Проверяет, есть ли уже даты в файле Excel
    
    Returns:
        bool: True если даты найдены, False если нет
    """
    # Проверяем столбец B (column=2)
    cell_value = ws.cell(row=1, column=2).value
    return is_date_string(cell_value)

def get_existing_dates(ws):
    """Получает список существующих дат в файле
    
    Returns:
        list: список дат в формате строк
    """
    existing_dates = []
    last_date_col = get_last_date_column(ws)
    
    for col in range(2, last_date_col + 1):
        cell_value = ws.cell(row=1, column=col).value
        if is_date_string(cell_value):
            if isinstance(cell_value, datetime):
                existing_dates.append(cell_value.strftime("%d.%m.%Y"))
            else:
                existing_dates.append(str(cell_value).strip())
    
    return existing_dates

def get_new_dates_needed(existing_dates, target_dates):
    """Определяет, какие даты нужно добавить
    
    Args:
        existing_dates: список существующих дат
        target_dates: список целевых дат для добавления
    
    Returns:
        list: список дат, которые нужно добавить
    """
    new_dates = []
    for target_date in target_dates:
        if target_date not in existing_dates:
            new_dates.append(target_date)
    
    return new_dates

def add_dates_and_grades_to_excel_files_in_folders(base_folder, months_to_process=None):
    """
    В каждой папке внутри base_folder ищет файлы Excel (xlsx), где есть ФИО студентов,
    и добавляет столбцы с датами указанных месяцев 2025 (только рабочие дни) и случайными оценками.
    
    Args:
        base_folder: путь к папке с группами
        months_to_process: список месяцев для обработки (например, [9, 10, 11, 12] для сентября-декабря)
    """
    if months_to_process is None:
        months_to_process = [9, 10, 11, 12]  # По умолчанию обрабатываем сентябрь-декабрь
    
    # Получаем рабочие дни для всех указанных месяцев
    all_working_days = []
    month_names = {9: "сентября", 10: "октября", 11: "ноября", 12: "декабря"}
    
    for month in months_to_process:
        month_days = get_working_days_for_month(2025, month)
        all_working_days.extend(month_days)
        print(f"Добавляем {len(month_days)} рабочих дней {month_names[month]} 2025")
    
    working_days = all_working_days
    
    # Стили для заголовков дат
    date_header_font = Font(bold=True, color="FFFFFF")
    date_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    date_header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Вероятность пропуска и стили для отметки "Н"
    absence_probability = 0.15  # 15% вероятность пропуска
    absence_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    absence_font = Font(bold=True, color="9C0006")
    
    # Распределение профилей студентов по типу оценок
    # only_5: только 5
    # only_4_5: только 4 и 5
    # no_4_5: только 2 и 3 (без 4 и 5)
    # mixed: любые 2-5
    prob_only_5 = 0.15
    prob_only_4_5 = 0.35
    prob_no_4_5 = 0.20
    # Остальное идёт на смешанный профиль
    
    for group_folder in os.listdir(base_folder):
        group_path = os.path.join(base_folder, group_folder)
        if os.path.isdir(group_path):
            for file in os.listdir(group_path):
                if file.endswith('.xlsx') and file != 'студенты.xlsx':  # Пропускаем файл со списком студентов
                    file_path = os.path.join(group_path, file)
                    try:
                        wb = load_workbook(file_path)
                        ws = wb.active
                        
                        # Проверяем, есть ли столбец "ФИО"
                        headers = [cell.value for cell in ws[1]]
                        if "ФИО" in headers:
                            # Проверяем, есть ли уже даты в файле
                            existing_dates = []
                            if has_existing_dates(ws):
                                existing_dates = get_existing_dates(ws)
                                print(f"  Найдены существующие даты: {len(existing_dates)}")
                                print(f"  Первая дата: {existing_dates[0] if existing_dates else 'Нет'}")
                                print(f"  Последняя дата: {existing_dates[-1] if existing_dates else 'Нет'}")
                            
                            # Определяем, какие даты нужно добавить
                            new_dates = get_new_dates_needed(existing_dates, working_days)
                            
                            if not new_dates:
                                print(f"  Все даты уже присутствуют в файле '{file}'")
                                continue
                            
                            print(f"  Добавляем {len(new_dates)} новых дат в файл '{file}'")
                            
                            # Находим последний столбец с датой
                            if existing_dates:
                                last_date_col = get_last_date_column(ws)
                                last_col = last_date_col
                            else:
                                last_col = 1  # Столбец A (ФИО)
                            
                            # Добавляем заголовки новых дат
                            for i, date in enumerate(new_dates):
                                col = last_col + 1 + i
                                cell = ws.cell(row=1, column=col, value=date)
                                cell.font = date_header_font
                                cell.fill = date_header_fill
                                cell.alignment = date_header_alignment
                            
                            # Добавляем оценки/пропуски с учётом профилей студентов
                            for row in range(2, ws.max_row + 1):
                                r = random.random()
                                if r < prob_only_5:
                                    profile = "only_5"
                                elif r < prob_only_5 + prob_only_4_5:
                                    profile = "only_4_5"
                                elif r < prob_only_5 + prob_only_4_5 + prob_no_4_5:
                                    profile = "no_4_5"
                                else:
                                    profile = "mixed"

                                # Добавляем оценки только для новых дат
                                for i, date in enumerate(new_dates):
                                    col = last_col + 1 + i
                                    # Случайный пропуск "Н" с заданной вероятностью, иначе оценка по профилю
                                    if random.random() < absence_probability:
                                        cell = ws.cell(row=row, column=col, value="Н")
                                        cell.font = absence_font
                                        cell.fill = absence_fill
                                        cell.alignment = Alignment(horizontal="center", vertical="center")
                                    else:
                                        if profile == "only_5":
                                            grade = 5
                                        elif profile == "only_4_5":
                                            grade = 5 if random.random() < 0.5 else 4
                                        elif profile == "no_4_5":
                                            grade = 3 if random.random() < 0.5 else 2
                                        else:
                                            grade = random.randint(2, 5)
                                        ws.cell(row=row, column=col, value=grade)
                            
                            # Автоматически подгоняем ширину столбцов
                            for column in ws.columns:
                                max_length = 0
                                column_letter = column[0].column_letter
                                for cell in column:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                adjusted_width = min(max_length + 2, 15)
                                ws.column_dimensions[column_letter].width = adjusted_width
                            
                            wb.save(file_path)
                            print(f"Файл '{file}' в папке '{group_folder}' обновлен.")
                            
                    except Exception as e:
                        print(f"Ошибка при обработке файла {file_path}: {e}")

def show_working_days_for_months(months_to_show):
    """Показывает рабочие дни для указанных месяцев"""
    month_names = {9: "сентября", 10: "октября", 11: "ноября", 12: "декабря"}
    
    for month in months_to_show:
        working_days = get_working_days_for_month(2025, month)
        print(f"\nРабочие дни {month_names[month]} 2025:")
        for i, day in enumerate(working_days, 1):
            print(f"{i:2d}. {day}")
        print(f"Всего рабочих дней {month_names[month]}: {len(working_days)}")

if __name__ == "__main__":
    # Используем папку с журналами, где находятся папки с группами
    base_folder = "Журналы/1 Курс"
    
    # Месяцы для обработки (сентябрь-декабрь 2025)
    months_to_process = [9, 10, 11, 12]
    
    print("Добавляем даты и оценки в файлы...")
    print("=" * 60)
    
    # Показываем рабочие дни для всех месяцев
    show_working_days_for_months(months_to_process)
    
    # Подсчитываем общее количество дней
    total_days = sum(len(get_working_days_for_month(2025, month)) for month in months_to_process)
    print(f"\nОбщее количество рабочих дней: {total_days}")
    print("-" * 60)
    
    # Запрашиваем подтверждение у пользователя
    response = input("Продолжить генерацию оценок? (y/n): ").strip().lower()
    if response in ['y', 'yes', 'да', 'д']:
        add_dates_and_grades_to_excel_files_in_folders(base_folder, months_to_process)
        print("\nГотово! Даты и оценки добавлены во все файлы по предметам.")
    else:
        print("Генерация отменена.")
