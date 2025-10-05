from faker import Faker
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Создаем экземпляр Faker для русского языка
fake = Faker('ru_RU')

def generate_person_data():
    """Генерирует от 20 до 30 случайных людей с полной информацией"""
    # Случайное количество людей от 20 до 30
    num_people = random.randint(20, 30)
    
    print(f"Генерируем {num_people} людей:")
    print("-" * 60)
    
    people = []
    for i in range(num_people):
        # Генерируем полную информацию о человеке
        person = {
            'фамилия': fake.last_name_male(),
            'имя': fake.first_name_male(),
            'отчество': fake.middle_name_male(),
        }
        people.append(person)
        
        # Выводим информацию в консоль
        print(f"{person['фамилия']} {person['имя']} {person['отчество']} ")
    
    print("-" * 60)
    print(f"Всего сгенерировано: {len(people)} людей")
    
    return people

def create_excel_file_with_groups(groups_list):
    """Создает Excel файл с отдельными листами для каждой группы"""
    # Создаем новую рабочую книгу
    wb = Workbook()
    
    # Удаляем стандартный лист
    wb.remove(wb.active)
    
    # Стили для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Заголовки столбцов
    headers = ['№', 'Фамилия', 'Имя', 'Отчество']
    
    for group_name in groups_list:
        # Создаем новый лист для группы
        ws = wb.create_sheet(title=group_name)
        
        # Заполняем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Генерируем студентов для этой группы (от 20 до 30 человек)
        num_students = random.randint(20, 30)
        students = []
        
        for i in range(num_students):
            student = {
                'номер': i + 1,
                'фамилия': fake.last_name(),
                'имя': fake.first_name(),
                'отчество': fake.middle_name()
            }
            students.append(student)
        
        # Заполняем данные студентов
        for row, student in enumerate(students, 2):
            ws.cell(row=row, column=1, value=student['номер'])
            ws.cell(row=row, column=2, value=student['фамилия'])
            ws.cell(row=row, column=3, value=student['имя'])
            ws.cell(row=row, column=4, value=student['отчество'])
        
        # Автоматически подгоняем ширину столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        print(f"Создан лист для группы: {group_name} ({num_students} студентов)")
    
    # Создаем имя файла с текущей датой и временем
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"список_групп_{timestamp}.xlsx"
    
    # Сохраняем файл
    wb.save(filename)
    print(f"\nExcel файл сохранен как: {filename}")
    
    return filename

group = [
    "ОДЛ-121",
    "ОДЛ-220(120)",
    "ОДЛ-316(216)",
    "ОДЛ-319(219)",
    "ТН-101",
    "ФК-201(101)",
    "Э-115",
    "Э-214(114)",
    "Э-313(213)",
    "ЭП-201(101)",
    "ПСА-301(201)",
    "ПСО-345(245)",
    "ЮР-148",
    "ЮР-149",
    "ЮР-246(146)",
    "ЮР-247(147)",
]

if __name__ == "__main__":
    # Создаем один Excel файл со всеми группами
    excel_filename = create_excel_file_with_groups(group)
    
    print(f"\nГотово! Создан Excel файл: {excel_filename}")
    print("Файл содержит:")
    print(f"- {len(group)} листов (по одному для каждой группы)")
    print("- На каждом листе список студентов с ФИО")
    print("- От 20 до 30 студентов в каждой группе")
    print("\nСписок групп:")
    for i, group_name in enumerate(group, 1):

        print(f"{i:2d}. {group_name}")
